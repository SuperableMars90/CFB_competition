"""
scripts/export_draft_guide.py
------------------------------
Exports a pre-draft schedule grid to Excel.

Rows: one FBS team per row, sorted P4 then G6, alphabetically by conference
      within each tier, then alphabetically by team within each conference.
      Independents appear last within their tier.

Columns: Team | Conference | Tier | Week 0 | Week 1 | ... | Week N

Cell values:
  - Home game:     "Opponent"   (out-of-conference, title case)
  - Away game:     "@Opponent"
  - In-conference: ALL CAPS (either prefix)
  - Neutral site:  italics
  - Bye / no game: "—"

Usage:
    PYTHONPATH=. python scripts/export_draft_guide.py --season 1
    PYTHONPATH=. python scripts/export_draft_guide.py --season 1 --output output/guide.xlsx
"""

import argparse
import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from db.connection import get_connection
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).parent.parent / 'output'

# -----------------------------------------------------------------------
# Sort helpers
# -----------------------------------------------------------------------

_TIER_RANK = {'P4': 0, 'G6': 1}


def _team_sort_key(team):
    tier_rank = _TIER_RANK.get(team['tier'], 2)
    is_ind = 'Independent' in (team['conf_name'] or '')
    conf = team['conf_name'] or ''
    name = team['name_display'] or ''
    return (tier_rank, is_ind, conf, name)


# -----------------------------------------------------------------------
# DB queries
# -----------------------------------------------------------------------

def _get_season_year(conn, season_id):
    cur = conn.cursor()
    cur.execute("SELECT year FROM seasons WHERE id = %s", (season_id,))
    row = cur.fetchone()
    cur.close()
    return row[0] if row else season_id


def get_fbs_teams(conn, season_id):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT t.id, t.name_display, t.conference_id,
               c.name  AS conf_name,
               c.abbreviation AS conf_abbrev,
               c.tier
        FROM teams t
        JOIN conferences c ON t.conference_id = c.id
        WHERE t.season_id = %s AND t.level = 'FBS'
    """, (season_id,))
    teams = cur.fetchall()
    cur.close()
    return sorted(teams, key=_team_sort_key)


def get_games(conn, season_id):
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT
            g.week,
            g.is_neutral,
            g.home_team_id,
            g.away_team_id,
            ht.name_display AS home_name,
            at.name_display AS away_name,
            ht.conference_id AS home_conf_id,
            at.conference_id AS away_conf_id
        FROM games g
        JOIN teams ht ON g.home_team_id = ht.id
        JOIN teams at ON g.away_team_id = at.id
        WHERE g.season_id = %s
        ORDER BY g.week
    """, (season_id,))
    games = cur.fetchall()
    cur.close()
    return games


# -----------------------------------------------------------------------
# Schedule pivot
# -----------------------------------------------------------------------

def build_schedule(fbs_teams, games):
    """
    Returns {team_id: {week: {opp_name, is_away, is_neutral, is_inconf}}}
    """
    fbs_ids = {t['id'] for t in fbs_teams}
    schedule = {t['id']: {} for t in fbs_teams}

    for game in games:
        home_id = game['home_team_id']
        away_id = game['away_team_id']
        week = game['week']
        is_neutral = bool(game['is_neutral'])
        is_inconf = (game['home_conf_id'] == game['away_conf_id'])

        if home_id in fbs_ids:
            schedule[home_id][week] = {
                'opp_name':  game['away_name'],
                'is_away':   False,
                'is_neutral': is_neutral,
                'is_inconf': is_inconf,
            }

        if away_id in fbs_ids:
            schedule[away_id][week] = {
                'opp_name':  game['home_name'],
                'is_away':   True,
                'is_neutral': is_neutral,
                'is_inconf': is_inconf,
            }

    return schedule


def _cell_text(game):
    """Format the opponent string according to display rules."""
    opp = game['opp_name']

    if game['is_inconf']:
        opp = opp.upper()

    # Neutral site gets no @ even if the team is listed as "away" in the DB
    if game['is_away'] and not game['is_neutral']:
        opp = '@' + opp

    return opp


# -----------------------------------------------------------------------
# Excel generation
# -----------------------------------------------------------------------

_FONT_NORMAL  = Font(name='Calibri', size=11)
_FONT_ITALIC  = Font(name='Calibri', size=11, italic=True)
_FONT_HEADER  = Font(name='Calibri', size=11, bold=True)
_FONT_SUBHDR  = Font(name='Calibri', size=10, bold=True)

_FILL_P4      = PatternFill('solid', fgColor='DCE6F1')   # soft blue
_FILL_G6      = PatternFill('solid', fgColor='EBF1DE')   # soft green
_FILL_HEADER  = PatternFill('solid', fgColor='4F81BD')
_FILL_WK0     = PatternFill('solid', fgColor='FFF2CC')   # yellow tint for week 0

_ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=False)
_ALIGN_LEFT   = Alignment(horizontal='left',   vertical='center')


def write_excel(fbs_teams, schedule, weeks, output_path, year):
    wb = Workbook()
    ws = wb.active
    ws.title = f'{year} Draft Guide'

    # ------------------------------------------------------------------
    # Header row
    # ------------------------------------------------------------------
    fixed_headers = ['Team', 'Conference', 'Tier']
    week_labels   = [('Wk 0' if w == 0 else f'Wk {w}') for w in weeks]
    all_headers   = fixed_headers + week_labels

    for col, label in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font      = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill      = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER

    # Highlight week 0 column header
    if 0 in weeks:
        wk0_col = weeks.index(0) + len(fixed_headers) + 1
        ws.cell(row=1, column=wk0_col).fill = PatternFill('solid', fgColor='C09000')

    # ------------------------------------------------------------------
    # Data rows
    # ------------------------------------------------------------------
    prev_tier = None

    for row_idx, team in enumerate(fbs_teams, 2):
        tier = team['tier']
        fill = _FILL_P4 if tier == 'P4' else (_FILL_G6 if tier == 'G6' else None)

        # Team name
        c = ws.cell(row=row_idx, column=1, value=team['name_display'])
        c.font = _FONT_NORMAL
        c.alignment = _ALIGN_LEFT
        if fill:
            c.fill = fill

        # Conference
        c = ws.cell(row=row_idx, column=2, value=team['conf_name'])
        c.font = _FONT_NORMAL
        c.alignment = _ALIGN_LEFT
        if fill:
            c.fill = fill

        # Tier
        c = ws.cell(row=row_idx, column=3, value=tier or '')
        c.font = _FONT_NORMAL
        c.alignment = _ALIGN_CENTER
        if fill:
            c.fill = fill

        # Week cells
        for col_offset, week in enumerate(weeks):
            col = len(fixed_headers) + col_offset + 1
            game = schedule[team['id']].get(week)

            if game is None:
                text     = '—'
                is_italic = False
            else:
                text      = _cell_text(game)
                is_italic = game['is_neutral']

            c = ws.cell(row=row_idx, column=col, value=text)
            c.font      = _FONT_ITALIC if is_italic else _FONT_NORMAL
            c.alignment = _ALIGN_CENTER
            if fill:
                c.fill = fill

    # ------------------------------------------------------------------
    # Column widths
    # ------------------------------------------------------------------
    ws.column_dimensions['A'].width = 24   # Team
    ws.column_dimensions['B'].width = 22   # Conference
    ws.column_dimensions['C'].width = 5    # Tier

    for col_offset in range(len(weeks)):
        letter = get_column_letter(len(fixed_headers) + col_offset + 1)
        ws.column_dimensions[letter].width = 20

    # Freeze panes: keep team/conf/tier visible while scrolling weeks
    ws.freeze_panes = 'D2'

    # Row height
    ws.row_dimensions[1].height = 18
    for row_idx in range(2, len(fbs_teams) + 2):
        ws.row_dimensions[row_idx].height = 15

    wb.save(output_path)


# -----------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='Export pre-draft schedule grid to Excel')
    parser.add_argument('--season', type=int, required=True, help='season_id in DB')
    parser.add_argument('--output', type=str, default='',
                        help='Output path (default: output/draft_guide_YEAR.xlsx)')
    args = parser.parse_args()

    conn = get_connection()
    year       = _get_season_year(conn, args.season)
    fbs_teams  = get_fbs_teams(conn, args.season)
    games      = get_games(conn, args.season)
    conn.close()

    schedule = build_schedule(fbs_teams, games)
    weeks    = sorted({g['week'] for g in games})

    output_path = args.output or str(OUTPUT_DIR / f'draft_guide_{year}.xlsx')
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    write_excel(fbs_teams, schedule, weeks, output_path, year)

    print(f"Draft guide written: {output_path}")
    print(f"  {len(fbs_teams)} FBS teams  |  weeks: {weeks[0]}–{weeks[-1]}")


if __name__ == '__main__':
    main()
