"""
scripts/export_draft_guide_fpi.py
----------------------------------
One-off variant of export_draft_guide.py that adds an FPI column, pulled live
from the CFBD /ratings/fpi endpoint, right in front of the Week 0 column.

Everything else (team sort order, schedule pivot, cell formatting) is unchanged
from export_draft_guide.py — this script imports those pieces directly rather
than duplicating them.

Usage:
    PYTHONPATH=. python scripts/export_draft_guide_fpi.py --season 8
    PYTHONPATH=. python scripts/export_draft_guide_fpi.py --season 8 --output output/guide.xlsx
"""

import argparse
import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from db.connection import get_connection
from scripts.cfbd_client import CFBDClient
from scripts.export_draft_guide import (
    _get_season_year,
    get_fbs_teams,
    get_games,
    build_schedule,
    _cell_text,
    _FONT_NORMAL,
    _FONT_ITALIC,
    _FILL_P4,
    _FILL_G6,
    _FILL_HEADER,
    _ALIGN_CENTER,
    _ALIGN_LEFT,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).parent.parent / 'output'


def get_fpi_ratings(year):
    """Returns {team_name: fpi_float} for the given year, matched by exact CFBD team name."""
    client = CFBDClient()
    ratings = client._get('/ratings/fpi', params={'year': year})
    return {r['team']: r['fpi'] for r in ratings if r.get('fpi') is not None}


def write_excel(fbs_teams, schedule, fpi_by_team, weeks, output_path, year):
    wb = Workbook()
    ws = wb.active
    ws.title = f'{year} Draft Guide'

    # ------------------------------------------------------------------
    # Header row
    # ------------------------------------------------------------------
    fixed_headers = ['Team', 'Conference', 'Tier', 'FPI']
    week_labels   = [('Wk 0' if w == 0 else f'Wk {w}') for w in weeks]
    all_headers   = fixed_headers + week_labels

    for col, label in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font      = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill      = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER

    # Highlight week 0 column header
    if 0 in weeks:
        wk0_col = weeks.index(0) + len(fixed_headers) + 1
        ws.cell(row=1, column=wk0_col).fill = PatternFill('solid', fgColor='C09000')

    # ------------------------------------------------------------------
    # Data rows
    # ------------------------------------------------------------------
    for row_idx, team in enumerate(fbs_teams, 2):
        tier = team['tier']
        fill = _FILL_P4 if tier == 'P4' else (_FILL_G6 if tier == 'G6' else None)

        c = ws.cell(row=row_idx, column=1, value=team['name_display'])
        c.font = _FONT_NORMAL
        c.alignment = _ALIGN_LEFT
        if fill:
            c.fill = fill

        c = ws.cell(row=row_idx, column=2, value=team['conf_name'])
        c.font = _FONT_NORMAL
        c.alignment = _ALIGN_LEFT
        if fill:
            c.fill = fill

        c = ws.cell(row=row_idx, column=3, value=tier or '')
        c.font = _FONT_NORMAL
        c.alignment = _ALIGN_CENTER
        if fill:
            c.fill = fill

        fpi = fpi_by_team.get(team['name_display'])
        c = ws.cell(row=row_idx, column=4, value=fpi)
        if fpi is not None:
            c.number_format = '0.0'
        c.font = _FONT_NORMAL
        c.alignment = _ALIGN_CENTER
        if fill:
            c.fill = fill

        for col_offset, week in enumerate(weeks):
            col = len(fixed_headers) + col_offset + 1
            game = schedule[team['id']].get(week)

            if game is None:
                text     = '—'
                is_italic = False
            else:
                text      = _cell_text(game)
                is_italic = game['is_neutral']

            c = ws.cell(row=row_idx, column=col, value=text)
            c.font      = _FONT_ITALIC if is_italic else _FONT_NORMAL
            c.alignment = _ALIGN_CENTER
            if fill:
                c.fill = fill

    # ------------------------------------------------------------------
    # Column widths
    # ------------------------------------------------------------------
    ws.column_dimensions['A'].width = 24   # Team
    ws.column_dimensions['B'].width = 22   # Conference
    ws.column_dimensions['C'].width = 5    # Tier
    ws.column_dimensions['D'].width = 8    # FPI

    for col_offset in range(len(weeks)):
        letter = get_column_letter(len(fixed_headers) + col_offset + 1)
        ws.column_dimensions[letter].width = 20

    # Freeze panes: keep team/conf/tier/FPI visible while scrolling weeks
    ws.freeze_panes = 'E2'

    ws.row_dimensions[1].height = 18
    for row_idx in range(2, len(fbs_teams) + 2):
        ws.row_dimensions[row_idx].height = 15

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description='Export pre-draft schedule grid to Excel, with an FPI column')
    parser.add_argument('--season', type=int, required=True, help='season_id in DB')
    parser.add_argument('--output', type=str, default='',
                        help='Output path (default: output/draft_guide_YEAR_fpi.xlsx)')
    args = parser.parse_args()

    conn = get_connection()
    year       = _get_season_year(conn, args.season)
    fbs_teams  = get_fbs_teams(conn, args.season)
    games      = get_games(conn, args.season)
    conn.close()

    schedule = build_schedule(fbs_teams, games)
    weeks    = sorted({g['week'] for g in games})
    fpi_by_team = get_fpi_ratings(year)

    missing = [t['name_display'] for t in fbs_teams if t['name_display'] not in fpi_by_team]
    if missing:
        print(f"Note: {len(missing)} team(s) have no CFBD FPI rating for {year}, left blank:")
        for name in missing:
            print(f"  - {name}")

    output_path = args.output or str(OUTPUT_DIR / f'draft_guide_{year}_fpi.xlsx')
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    write_excel(fbs_teams, schedule, fpi_by_team, weeks, output_path, year)

    print(f"Draft guide (with FPI) written: {output_path}")
    print(f"  {len(fbs_teams)} FBS teams  |  weeks: {weeks[0]}-{weeks[-1]}")


if __name__ == '__main__':
    main()
